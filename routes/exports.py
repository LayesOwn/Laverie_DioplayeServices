"""
Blueprint Exports — CSV, Excel, PDF reçu
"""
import io
import csv
from datetime import date
from flask import Blueprint, Response, send_file, request, flash, redirect, url_for
from flask_login import login_required
from extensions import db
from models import Transaction, DepenseInterne, Client

exports_bp = Blueprint("exports", __name__)


# ── Export CSV Transactions ───────────────────────
@exports_bp.route("/transactions/csv")
@login_required
def transactions_csv():
    transactions = (
        Transaction.query
        .filter(Transaction.deleted_at == None)  # noqa: E711
        .order_by(Transaction.date_transaction.desc())
        .all()
    )
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["ID", "Date", "Client", "Service", "Quantité",
                     "Total (XOF)", "Payé (XOF)", "Reste dû (XOF)", "Notes"])
    for t in transactions:
        writer.writerow([
            t.id,
            t.date_transaction.strftime("%d/%m/%Y"),
            t.client.nom,
            t.service.nom,
            t.quantite,
            t.total,
            t.total_paye,
            t.solde_restant,
            t.notes or "",
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=transactions_{date.today()}.csv"}
    )


# ── Export CSV Dépenses ───────────────────────────
@exports_bp.route("/depenses/csv")
@login_required
def depenses_csv():
    depenses = DepenseInterne.query.order_by(DepenseInterne.date_depense.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["ID", "Date", "Libellé", "Type", "Catégorie", "Montant (XOF)", "Notes"])
    for d in depenses:
        writer.writerow([
            d.id, d.date_depense.strftime("%d/%m/%Y"),
            d.libelle, d.type_depense, d.categorie or "",
            d.montant, d.notes or "",
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=depenses_{date.today()}.csv"}
    )


# ── Export Excel Transactions ─────────────────────
@exports_bp.route("/transactions/excel")
@login_required
def transactions_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("openpyxl non installé.", "danger")
        return redirect(url_for("transactions.index"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # En-tête
    headers = ["ID", "Date", "Client", "Service", "Qté", "Total", "Payé", "Reste dû", "Notes"]
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    transactions = (
        Transaction.query
        .filter(Transaction.deleted_at == None)  # noqa: E711
        .order_by(Transaction.date_transaction.desc())
        .all()
    )
    for row, t in enumerate(transactions, 2):
        ws.append([
            t.id,
            t.date_transaction.strftime("%d/%m/%Y"),
            t.client.nom,
            t.service.nom,
            t.quantite,
            t.total,
            t.total_paye,
            t.solde_restant,
            t.notes or "",
        ])

    # Largeur automatique
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"transactions_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Reçu PDF ─────────────────────────────────────
@exports_bp.route("/recu/<int:tx_id>/pdf")
@login_required
def recu_pdf(tx_id: int):
    try:
        from reportlab.lib.pagesizes import A5
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        flash("reportlab non installé.", "danger")
        return redirect(url_for("transactions.index"))

    tx = db.get_or_404(Transaction, tx_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("titre", parent=styles["Title"], fontSize=16,
                                 textColor=colors.HexColor("#0d6efd"))
    sous_titre = ParagraphStyle("sous", parent=styles["Normal"], fontSize=10,
                                textColor=colors.grey, alignment=1)

    story = [
        Paragraph("DIOPLAYE SERVICES", titre_style),
        Paragraph("Reçu de transaction", sous_titre),
        Spacer(1, 0.5*cm),
    ]

    data = [
        ["Reçu N°", f"#{tx.id:04d}"],
        ["Date", tx.date_transaction.strftime("%d/%m/%Y")],
        ["Client", tx.client.nom],
        ["Téléphone", tx.client.telephone or "—"],
        ["Service", tx.service.nom],
        ["Quantité", str(tx.quantite)],
        ["Prix unitaire", f"{tx.service.prix_unitaire:,.0f} XOF"],
        ["TOTAL", f"{tx.total:,.0f} XOF"],
        ["Montant payé", f"{tx.total_paye:,.0f} XOF"],
        ["Reste dû", f"{tx.solde_restant:,.0f} XOF"],
    ]

    table = Table(data, colWidths=[4*cm, 8*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f4ff")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 7), (0, 7), "Helvetica-Bold"),
        ("FONTNAME", (1, 7), (1, 7), "Helvetica-Bold"),
        ("TEXTCOLOR", (1, 7), (1, 7), colors.HexColor("#198754")),
        ("TEXTCOLOR", (1, 9), (1, 9), colors.HexColor("#dc3545")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(table)
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("Merci pour votre confiance !", sous_titre))

    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=False,
                     download_name=f"recu_{tx.id:04d}.pdf",
                     mimetype="application/pdf")